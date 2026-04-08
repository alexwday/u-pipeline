"""XLSX processor — reads workbooks via openpyxl, extracts to markdown."""

import base64
import hashlib
import json
import logging
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from datetime import date, datetime
from itertools import zip_longest
from pathlib import Path
from typing import Any

import openai
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries, range_to_tuple
from openpyxl.worksheet.worksheet import Worksheet

from ...utils.llm_connector import LLMClient
from ...utils.config_setup import (
    get_extraction_page_workers,
    get_xlsx_vision_max_retries,
    get_xlsx_vision_retry_delay,
)
from ...utils.file_types import ExtractionResult, PageResult
from ...utils.prompt_loader import load_prompt

logger = logging.getLogger(__name__)
_PROMPTS_DIR = Path(__file__).resolve().parent / "prompts"

RETRYABLE_ERRORS = (
    openai.RateLimitError,
    openai.APITimeoutError,
    openai.APIConnectionError,
    openai.InternalServerError,
)

_PARSE_RETRYABLE_ERRORS = RETRYABLE_ERRORS + (ValueError,)

_VISUAL_SERIES_POINT_LIMIT = 12
_IMAGE_MEDIA_TYPES = {
    "gif": "image/gif",
    "jpeg": "image/jpeg",
    "jpg": "image/jpeg",
    "png": "image/png",
}
_VISUAL_DESCRIPTION_ERRORS = RETRYABLE_ERRORS + (
    RuntimeError,
    ValueError,
)


@dataclass
class SheetData:
    """Pre-read sheet data for thread-safe parallel processing.

    All openpyxl-dependent data is extracted during the sequential
    phase so the parallel phase never touches workbook objects.

    Params:
        title: Worksheet tab name
        page_number: 1-indexed position among workbook sheets
        is_chartsheet: True when the tab is a dedicated chartsheet
        cell_data: Grid structure from _collect_populated_cells,
            or None for empty / chart-only sheets
        chart_descriptions: Pre-formatted chart markdown strings
        image_data: Pre-loaded image bytes with metadata dicts
    """

    title: str
    page_number: int
    is_chartsheet: bool
    cell_data: dict[str, Any] | None
    chart_descriptions: list[str]
    image_data: list[dict[str, Any]] = field(
        default_factory=list,
    )


@dataclass
class _InflightImageDescription:
    """Track one in-flight deduplicated image description."""

    event: threading.Event = field(default_factory=threading.Event)
    description: str = ""
    error: Exception | None = None


@dataclass
class _ImageDescriptionState:
    """Hold shared cache state for deduplicated image descriptions."""

    descriptions: dict[str, str] = field(default_factory=dict)
    in_flight: dict[str, _InflightImageDescription] = field(
        default_factory=dict,
    )
    lock: threading.Lock = field(default_factory=threading.Lock)


def _preread_image_data(image: Any) -> dict[str, Any] | None:
    """Pre-read image bytes and metadata from an openpyxl image.

    Params:
        image: openpyxl drawing image object

    Returns:
        dict with bytes, format, and anchor — or None when
        the image object cannot expose raw bytes.
    """
    if not hasattr(image, "_data"):
        return None
    loader = getattr(image, "_data")
    return {
        "bytes": loader(),
        "format": str(getattr(image, "format", "png")).lower(),
        "anchor": getattr(image, "anchor", None),
    }


# -----------------------------------------------------------------
# Cell normalization
# -----------------------------------------------------------------


def _normalize_cell_value(
    value: Any,
    number_format: str = "",
) -> str:
    """Normalize a worksheet cell value into stable text.

    Params:
        value: Raw cell value from openpyxl
        number_format: Excel number format string for the cell

    Returns:
        Cleaned string representation
    """
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, datetime):
        if value.hour == 0 and value.minute == 0 and value.second == 0:
            return value.strftime("%Y-%m-%d")
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float):
        if number_format and "%" in number_format:
            return format(value * 100, ".10g") + "%"
        return format(value, ".15g")
    if isinstance(value, int):
        return str(value)
    return str(value).replace("\r\n", "\n").replace("\r", "\n")


def _escape_table_text(value: str) -> str:
    """Escape markdown table delimiters in cell text."""
    return value.replace("|", "\\|").replace("\n", "<br>")


# -----------------------------------------------------------------
# Workbook I/O
# -----------------------------------------------------------------


def _open_workbooks(file_path: str) -> tuple[Any, Any]:
    """Open formula and cached-value workbooks from the same file."""
    name = Path(file_path).name
    try:
        workbook: Any = load_workbook(filename=file_path, data_only=False)
    except (OSError, ValueError) as exc:
        raise RuntimeError(f"Failed to open XLSX '{name}': {exc}") from exc

    success = False
    try:
        try:
            cached_workbook: Any = load_workbook(
                filename=file_path, data_only=True
            )
        except (OSError, ValueError) as exc:
            raise RuntimeError(f"Failed to open XLSX '{name}': {exc}") from exc
        success = True
        return workbook, cached_workbook
    finally:
        if not success:
            workbook.close()


def _build_cached_values(
    cached_sheet: Worksheet,
) -> dict[tuple[int, int], Any]:
    """Build a lookup of cached display values from a data_only sheet."""
    values: dict[tuple[int, int], Any] = {}
    for row in cached_sheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                values[(cell.row, cell.column)] = cell.value
    return values


def _is_chartsheet(sheet: Any) -> bool:
    """Return True when the workbook tab is a dedicated chartsheet."""
    return type(sheet).__name__ == "Chartsheet"


# -----------------------------------------------------------------
# Grid serialization
# -----------------------------------------------------------------


def _collect_populated_cells(
    sheet: Worksheet,
    cached_values: dict[tuple[int, int], Any],
) -> dict[str, Any] | None:
    """Collect non-empty worksheet cells into a grid structure.

    Returns None for sheets with no populated cells.
    """
    min_row = 0
    min_col = 0
    max_row = 0
    max_col = 0
    rows: dict[int, dict[int, str]] = {}
    populated_columns: set[int] = set()

    for row in sheet.iter_rows():
        populated_cells_in_row: dict[int, str] = {}
        for cell in row:
            value = _normalize_cell_value(
                cached_values.get((cell.row, cell.column), cell.value),
                number_format=cell.number_format or "",
            )
            if not value:
                continue
            if min_row == 0 or cell.row < min_row:
                min_row = cell.row
            if min_col == 0 or cell.column < min_col:
                min_col = cell.column
            max_row = max(max_row, cell.row)
            max_col = max(max_col, cell.column)
            populated_cells_in_row[cell.column] = value
            populated_columns.add(cell.column)
        if populated_cells_in_row:
            rows[row[0].row] = populated_cells_in_row

    if min_row == 0:
        return None
    return {
        "rows": rows,
        "columns": sorted(populated_columns),
    }


def _serialize_sheet(
    sheet_name: str,
    cell_data: dict[str, Any],
) -> str:
    """Serialize worksheet cells into a markdown table.

    Params:
        sheet_name: Worksheet name for the heading
        cell_data: Output from _collect_populated_cells

    Returns:
        str — markdown content with heading and table
    """
    columns = cell_data["columns"]
    header_cells = [get_column_letter(col) for col in columns]
    lines = [
        f"# Sheet: {sheet_name}",
        "",
        f"| Row | {' | '.join(header_cells)} |",
        f"| {' | '.join(['---'] * (len(header_cells) + 1))} |",
    ]
    for row_number in sorted(cell_data["rows"]):
        row_values = cell_data["rows"][row_number]
        rendered_row = [
            _escape_table_text(row_values.get(col, "")) for col in columns
        ]
        lines.append(f"| {row_number} | {' | '.join(rendered_row)} |")
    return "\n".join(lines)


# -----------------------------------------------------------------
# Chart metadata extraction (programmatic)
# -----------------------------------------------------------------


def _extract_title_text(title: Any) -> str:
    """Extract plain text from an openpyxl title-like object."""
    if title is None:
        return ""
    if isinstance(title, str):
        return title.strip()

    text_runs: list[str] = []
    text = getattr(title, "tx", None)
    rich_text = getattr(text, "rich", None)
    paragraphs = getattr(rich_text, "p", [])
    for paragraph in paragraphs:
        for run in getattr(paragraph, "r", []) or []:
            run_text = getattr(run, "t", "")
            if run_text:
                text_runs.append(str(run_text))
        for fld_item in getattr(paragraph, "fld", []) or []:
            field_text = getattr(fld_item, "t", "")
            if field_text:
                text_runs.append(str(field_text))
    return " ".join(part.strip() for part in text_runs if part).strip()


def _extract_chart_title(chart: Any) -> str:
    """Extract plain text from an openpyxl chart title."""
    return _extract_title_text(getattr(chart, "title", None))


def _extract_anchor_cell(anchor: Any) -> str:
    """Extract cell anchor string from a drawing anchor."""
    marker = getattr(anchor, "_from", None)
    if marker is None and hasattr(anchor, "from_"):
        marker = getattr(anchor, "from_", None)
    if marker is not None:
        row = getattr(marker, "row", None)
        column = getattr(marker, "col", None)
        if isinstance(row, int) and isinstance(column, int):
            return f"{get_column_letter(column + 1)}{row + 1}"

    if isinstance(anchor, str):
        min_col, min_row, _max_col, _max_row = range_boundaries(anchor)
        return f"{get_column_letter(min_col)}{min_row}"

    return "unknown"


def _load_reference_values(
    workbook: Any,
    cached_workbook: Any,
    reference: str,
) -> list[str]:
    """Resolve an Excel reference formula into normalized values."""
    if not reference:
        return []
    try:
        sheet_name, boundaries = range_to_tuple(reference)
    except ValueError:
        return []

    min_col, min_row, max_col, max_row = boundaries
    try:
        formula_sheet = workbook[sheet_name]
    except KeyError:
        return []

    cached_sheet = None
    if sheet_name in getattr(cached_workbook, "sheetnames", []):
        cached_sheet = cached_workbook[sheet_name]

    values: list[str] = []
    for row_number in range(min_row, max_row + 1):
        for column_number in range(min_col, max_col + 1):
            formula_cell = formula_sheet.cell(
                row=row_number, column=column_number
            )
            raw_value = (
                cached_sheet.cell(row=row_number, column=column_number).value
                if cached_sheet is not None
                else formula_cell.value
            )
            values.append(
                _normalize_cell_value(
                    raw_value,
                    number_format=formula_cell.number_format or "",
                )
            )
    return values


def _extract_series_name(
    series: Any,
    workbook: Any,
    cached_workbook: Any,
    series_index: int,
) -> str:
    """Extract a human-readable chart series name."""
    text_source = getattr(series, "tx", None)
    literal_value = getattr(text_source, "v", None)
    if isinstance(literal_value, str) and literal_value.strip():
        return literal_value.strip()

    string_reference = getattr(text_source, "strRef", None)
    formula = getattr(string_reference, "f", "")
    values = _load_reference_values(workbook, cached_workbook, formula)
    for value in values:
        if value:
            return value
    return f"Series {series_index}"


def _extract_series_points(
    series: Any,
    workbook: Any,
    cached_workbook: Any,
) -> list[str]:
    """Extract paired category/value points for a chart series."""
    category_source = getattr(series, "cat", None)
    category_reference = getattr(category_source, "strRef", None)
    if category_reference is None:
        category_reference = getattr(category_source, "numRef", None)
    category_formula = getattr(category_reference, "f", "")
    value_source = getattr(series, "val", None)
    value_reference = getattr(value_source, "numRef", None)
    value_formula = getattr(value_reference, "f", "")

    categories = _load_reference_values(
        workbook, cached_workbook, category_formula
    )
    values = _load_reference_values(workbook, cached_workbook, value_formula)
    points: list[str] = []
    for index, (category, value) in enumerate(
        zip_longest(categories, values, fillvalue=""),
        start=1,
    ):
        if not category and not value:
            continue
        label = category or f"Point {index}"
        points.append(f"{label}: {value}")
        if len(points) >= _VISUAL_SERIES_POINT_LIMIT:
            break
    return points


def _format_chart_markdown(
    sheet_name: str,
    chart: Any,
    workbook: Any,
    cached_workbook: Any,
) -> str:
    """Format one Excel chart as markdown from programmatic metadata.

    Params:
        sheet_name: Worksheet name for context
        chart: Openpyxl chart object
        workbook: Formula workbook handle
        cached_workbook: Cached-value workbook handle

    Returns:
        str — markdown description of the chart
    """
    anchor_cell = _extract_anchor_cell(getattr(chart, "anchor", None))
    chart_type = type(chart).__name__
    chart_title = _extract_chart_title(chart) or "Untitled Chart"
    x_title = _extract_title_text(getattr(chart.x_axis, "title", None))
    y_title = _extract_title_text(getattr(chart.y_axis, "title", None))

    header = f'> [Chart]: {chart_type} — "{chart_title}"'
    context_parts = []
    if x_title:
        context_parts.append(f"X-axis: {x_title}")
    if y_title:
        context_parts.append(f"Y-axis: {y_title}")
    context_parts.append(f"Sheet: {sheet_name}, Cell: {anchor_cell}")

    lines = [header]
    if context_parts:
        lines.append(f"> {', '.join(context_parts)}")

    series_list = getattr(chart, "ser", [])
    for series_index, series in enumerate(series_list, start=1):
        name = _extract_series_name(
            series, workbook, cached_workbook, series_index
        )
        points = _extract_series_points(series, workbook, cached_workbook)
        if points:
            lines.append(f"> Series \"{name}\": {', '.join(points)}")
        else:
            lines.append(f'> Series "{name}": no data points recovered')

    return "\n".join(lines)


# -----------------------------------------------------------------
# LLM visual extraction
# -----------------------------------------------------------------


def _parse_visual_response(response: dict) -> str:
    """Extract content string from a visual extraction response."""
    choices = response.get("choices")
    if not isinstance(choices, list) or not choices:
        raise ValueError("LLM response missing choices")

    finish_reason = choices[0].get("finish_reason", "")
    message = choices[0].get("message")
    if not isinstance(message, dict):
        raise ValueError("LLM response missing message payload")

    tool_calls = message.get("tool_calls")
    if not isinstance(tool_calls, list) or not tool_calls:
        content_preview = str(message.get("content", ""))[:200]
        raise ValueError(
            f"LLM response missing tool calls "
            f"(finish_reason={finish_reason}, "
            f"content={content_preview!r})"
        )

    function_data = tool_calls[0].get("function")
    if not isinstance(function_data, dict):
        raise ValueError("LLM response missing function payload")

    arguments = function_data.get("arguments")
    if not isinstance(arguments, str):
        raise ValueError("LLM response missing function arguments")

    parsed = json.loads(arguments)
    if not isinstance(parsed, dict):
        raise ValueError("LLM tool arguments must decode to an object")

    content = parsed.get("content")
    if not isinstance(content, str) or not content.strip():
        raise ValueError("LLM returned empty or missing content")
    return content


def _llm_call_with_retry(
    llm: LLMClient,
    messages: list,
    prompt: dict[str, Any],
    context: str,
) -> str:
    """Make an LLM call with transient error retries.

    Params:
        llm: LLMClient instance
        messages: Message list for the API call
        prompt: Loaded prompt dict (for stage, tools, tool_choice)
        context: Log label

    Returns:
        str — parsed content from the tool response
    """
    max_retries = get_xlsx_vision_max_retries()
    retry_delay = get_xlsx_vision_retry_delay()

    for attempt in range(1, max_retries + 1):
        try:
            response = llm.call(
                messages=messages,
                stage=prompt["stage"],
                tools=prompt["tools"],
                tool_choice=prompt.get("tool_choice", "required"),
                context=context,
            )
            return _parse_visual_response(response)
        except _PARSE_RETRYABLE_ERRORS as exc:
            if attempt == max_retries:
                logger.error(
                    "%s failed after %d retries: %s",
                    context,
                    max_retries,
                    exc,
                )
                raise
            wait = retry_delay * attempt
            logger.warning(
                "%s retry %d/%d after %.1fs: %s",
                context,
                attempt,
                max_retries,
                wait,
                exc,
            )
            time.sleep(wait)
    raise RuntimeError(f"{context} exited retry loop without a response")


def _describe_image(
    llm: LLMClient,
    prompt: dict[str, Any],
    sheet_name: str,
    image: Any,
    context: str,
    image_cache: dict[str, str],
) -> str:
    """Describe one embedded image via LLM vision, reusing cached descriptions.

    Params:
        llm: LLMClient instance
        prompt: Loaded prompt dict
        sheet_name: Current worksheet title
        image: openpyxl image object
        context: Log label
        image_cache: SHA-256 -> description cache for deduplication
    """
    if not hasattr(image, "_data"):
        raise RuntimeError(
            f"Embedded image bytes unavailable in '{sheet_name}'"
        )

    anchor_cell = _extract_anchor_cell(getattr(image, "anchor", None))
    loader = getattr(image, "_data")
    image_bytes = loader()
    image_hash = hashlib.sha256(image_bytes).hexdigest()

    if image_hash in image_cache:
        logger.info("%s — deduplicated (reusing cached description)", context)
        return image_cache[image_hash]

    b64 = base64.b64encode(image_bytes).decode()
    image_format = str(getattr(image, "format", "png")).lower()
    media_type = {
        "gif": "image/gif",
        "jpeg": "image/jpeg",
        "jpg": "image/jpeg",
        "png": "image/png",
    }.get(image_format, "image/png")

    user_text = (
        f"{prompt['user_prompt']}\n\n"
        "## Image metadata\n"
        f"- Sheet: {sheet_name}\n"
        f"- Anchor cell: {anchor_cell}\n"
        "- Type: embedded image"
    )

    messages = []
    if prompt.get("system_prompt"):
        messages.append({"role": "system", "content": prompt["system_prompt"]})
    messages.append(
        {
            "role": "user",
            "content": [
                {
                    "type": "image_url",
                    "image_url": {
                        "url": f"data:{media_type};base64,{b64}",
                        "detail": "high",
                    },
                },
                {"type": "text", "text": user_text},
            ],
        },
    )
    description = _llm_call_with_retry(llm, messages, prompt, context)
    image_cache[image_hash] = description
    return description


def _format_all_charts(
    sheet: Any,
    workbook: Any,
    cached_workbook: Any,
) -> list[str]:
    """Format all charts on a sheet as markdown (programmatic)."""
    return [
        _format_chart_markdown(sheet.title, chart, workbook, cached_workbook)
        for chart in getattr(sheet, "_charts", [])
    ]


def _describe_all_images(
    llm: LLMClient,
    prompt: dict[str, Any],
    sheet: Any,
    file_label: str,
    image_cache: dict[str, str],
) -> list[str]:
    """Describe all embedded images on a sheet via LLM vision."""
    descriptions: list[str] = []
    sheet_name = sheet.title
    for idx, image in enumerate(getattr(sheet, "_images", []), start=1):
        context = f"{file_label} '{sheet_name}' image {idx}"
        descriptions.append(
            _describe_image(
                llm, prompt, sheet_name, image, context, image_cache
            )
        )
    return descriptions


def _get_image_media_type(image_format: str) -> str:
    """Return the prompt media type for an image format.

    Params:
        image_format: Image format name, such as png or jpeg

    Returns:
        str — media type string for the vision prompt
    """
    return _IMAGE_MEDIA_TYPES.get(image_format, "image/png")


def _build_preloaded_image_messages(
    prompt: dict[str, Any],
    sheet_name: str,
    image_data: dict[str, Any],
) -> list[dict[str, Any]]:
    """Build LLM messages for a pre-loaded sheet image.

    Params:
        prompt: Loaded prompt dict
        sheet_name: Current worksheet title
        image_data: Pre-loaded dict with bytes, format, and anchor

    Returns:
        list[dict[str, Any]] — chat messages for the vision call
    """
    anchor_cell = _extract_anchor_cell(image_data["anchor"])
    encoded_image = base64.b64encode(image_data["bytes"]).decode()
    media_type = _get_image_media_type(image_data["format"])
    user_text = (
        f"{prompt['user_prompt']}\n\n"
        "## Image metadata\n"
        f"- Sheet: {sheet_name}\n"
        f"- Anchor cell: {anchor_cell}\n"
        "- Type: embedded image"
    )

    messages: list[dict[str, Any]] = []
    if prompt.get("system_prompt"):
        messages.append({"role": "system", "content": prompt["system_prompt"]})
    messages.append(
        {
            "role": "user",
            "content": [
                {
                    "type": "image_url",
                    "image_url": {
                        "url": f"data:{media_type};base64,{encoded_image}",
                        "detail": "high",
                    },
                },
                {"type": "text", "text": user_text},
            ],
        },
    )
    return messages


def _claim_preloaded_image_description(
    cache_state: _ImageDescriptionState,
    image_hash: str,
    context: str,
) -> tuple[str | None, _InflightImageDescription | None, bool]:
    """Claim or join a deduplicated image-description request.

    Params:
        cache_state: Shared deduplication state
        image_hash: SHA-256 of the image bytes
        context: Log label

    Returns:
        tuple of cached description, in-flight state,
        and ownership flag
    """
    with cache_state.lock:
        cached = cache_state.descriptions.get(image_hash)
        if cached is not None:
            logger.info(
                "%s — deduplicated (reusing cached description)",
                context,
            )
            return cached, None, False

        flight = cache_state.in_flight.get(image_hash)
        if flight is None:
            flight = _InflightImageDescription()
            cache_state.in_flight[image_hash] = flight
            return None, flight, True

        return None, flight, False


def _wait_for_preloaded_image_description(
    flight: _InflightImageDescription,
    context: str,
) -> str:
    """Wait for another worker to finish a deduplicated description request.

    Params:
        flight: In-flight description state owned by another worker
        context: Log label

    Returns:
        str — the shared description
    """
    flight.event.wait()
    if flight.error is not None:
        raise RuntimeError(
            f"{context} failed while waiting for a deduplicated "
            f"image description: {flight.error}"
        ) from flight.error

    logger.info(
        "%s — deduplicated (reusing in-flight description)",
        context,
    )
    return flight.description


def _finalize_preloaded_image_description(
    cache_state: _ImageDescriptionState,
    image_hash: str,
    flight: _InflightImageDescription,
    context: str,
    description: str,
    completed: bool,
) -> None:
    """Publish the result of an owned image-description request.

    Params:
        cache_state: Shared deduplication state
        image_hash: SHA-256 of the image bytes
        flight: In-flight state to resolve
        context: Log label
        description: Final description, if produced
        completed: True when the owner finished successfully
    """
    with cache_state.lock:
        if completed:
            cache_state.descriptions[image_hash] = description
            flight.description = description
        elif flight.error is None:
            flight.error = RuntimeError(
                f"{context} failed before caching a deduplicated "
                f"image description"
            )
        cache_state.in_flight.pop(image_hash, None)
        flight.event.set()


def _run_preloaded_image_description(
    llm: LLMClient,
    messages: list[dict[str, Any]],
    prompt: dict[str, Any],
    context: str,
    image_hash: str,
    cache_state: _ImageDescriptionState,
    flight: _InflightImageDescription,
) -> str:
    """Execute and publish one owned deduplicated image-description request.

    Params:
        llm: LLMClient instance
        messages: Chat messages for the vision call
        prompt: Loaded prompt dict
        context: Log label
        image_hash: SHA-256 of the image bytes
        cache_state: Shared deduplication state
        flight: In-flight state owned by this worker

    Returns:
        str — markdown description of the image
    """
    description = ""
    completed = False
    try:
        description = _llm_call_with_retry(llm, messages, prompt, context)
        completed = True
        return description
    except _VISUAL_DESCRIPTION_ERRORS as exc:
        with cache_state.lock:
            flight.error = exc
        raise
    finally:
        _finalize_preloaded_image_description(
            cache_state,
            image_hash,
            flight,
            context,
            description,
            completed,
        )


def _describe_preloaded_image(
    llm: LLMClient,
    prompt: dict[str, Any],
    sheet_name: str,
    image_data: dict[str, Any],
    context: str,
    cache_state: _ImageDescriptionState | None = None,
) -> str:
    """Describe a pre-loaded image via LLM with thread-safe caching.

    Works with pre-extracted bytes so no openpyxl access is needed.
    The shared cache state protects concurrent deduplicated requests.

    Params:
        llm: LLMClient instance
        prompt: Loaded prompt dict
        sheet_name: Current worksheet title
        image_data: Pre-loaded dict with bytes, format, anchor
        context: Log label
        cache_state: Optional shared cache/in-flight dedup state

    Returns:
        str — markdown description of the image
    """
    state = _ImageDescriptionState() if cache_state is None else cache_state
    image_hash = hashlib.sha256(image_data["bytes"]).hexdigest()
    cached, flight, should_describe = _claim_preloaded_image_description(
        state,
        image_hash,
        context,
    )
    if cached is not None:
        return cached
    if not should_describe:
        return _wait_for_preloaded_image_description(flight, context)
    messages = _build_preloaded_image_messages(
        prompt,
        sheet_name,
        image_data,
    )
    return _run_preloaded_image_description(
        llm,
        messages,
        prompt,
        context,
        image_hash,
        state,
        flight,
    )


def _process_single_sheet(
    sheet_data: SheetData,
    llm: LLMClient,
    prompt: dict[str, Any] | None,
    file_label: str,
    total_sheets: int,
    cache_state: _ImageDescriptionState | None = None,
) -> PageResult:
    """Process one pre-read sheet into a PageResult.

    Serializes cells to markdown, describes images via LLM (using
    pre-loaded bytes), and assembles final content. Safe to call
    from multiple threads — no openpyxl access occurs.

    Params:
        sheet_data: Pre-read sheet data from the sequential phase
        llm: LLMClient instance
        prompt: Visual extraction prompt dict, or None
        file_label: Workbook filename for log context
        total_sheets: Total number of sheets in the workbook
        cache_state: Optional shared image-description dedup state

    Returns:
        PageResult with extracted markdown content
    """
    image_descriptions: list[str] = []
    if sheet_data.image_data and prompt is not None:
        for idx, img in enumerate(sheet_data.image_data, start=1):
            context = f"{file_label} '{sheet_data.title}' image {idx}"
            image_descriptions.append(
                _describe_preloaded_image(
                    llm,
                    prompt,
                    sheet_data.title,
                    img,
                    context,
                    cache_state,
                )
            )

    visual_descriptions = sheet_data.chart_descriptions + image_descriptions

    if sheet_data.cell_data is None and not visual_descriptions:
        content = (
            f"# Sheet: {sheet_data.title}\n\n" "This sheet contains no data."
        )
    elif sheet_data.cell_data is None:
        parts = [f"# Sheet: {sheet_data.title}"]
        parts.extend(visual_descriptions)
        content = "\n\n".join(parts)
    else:
        grid = _serialize_sheet(sheet_data.title, sheet_data.cell_data)
        if visual_descriptions:
            parts = [grid]
            parts.extend(visual_descriptions)
            content = "\n\n".join(parts)
        else:
            content = grid

    logger.info(
        "%s sheet %d/%d '%s' extracted (%d chars)",
        file_label,
        sheet_data.page_number,
        total_sheets,
        sheet_data.title,
        len(content),
    )

    return PageResult(
        page_number=sheet_data.page_number,
        raw_content=content,
    )


# -----------------------------------------------------------------
# Orchestrator
# -----------------------------------------------------------------


def _preread_all_sheets(
    workbook: Any,
    cached_workbook: Any,
) -> list[SheetData]:
    """Pre-read all sheet data from openpyxl into plain structures.

    This sequential phase extracts every openpyxl-dependent value
    so the parallel phase never touches workbook objects.

    Params:
        workbook: Formula workbook handle
        cached_workbook: Cached-value workbook handle

    Returns:
        list[SheetData] — one entry per workbook sheet
    """
    sheets = [workbook[name] for name in workbook.sheetnames]
    preread: list[SheetData] = []

    for page_number, sheet in enumerate(sheets, start=1):
        chartsheet = _is_chartsheet(sheet)

        if chartsheet:
            cached_values: dict[tuple[int, int], Any] = {}
        else:
            cached_values = _build_cached_values(cached_workbook[sheet.title])

        chart_descriptions = _format_all_charts(
            sheet, workbook, cached_workbook
        )

        images: list[dict[str, Any]] = []
        for image in getattr(sheet, "_images", []):
            data = _preread_image_data(image)
            if data is not None:
                images.append(data)

        cell_data = (
            None
            if chartsheet
            else _collect_populated_cells(sheet, cached_values)
        )

        preread.append(
            SheetData(
                title=sheet.title,
                page_number=page_number,
                is_chartsheet=chartsheet,
                cell_data=cell_data,
                chart_descriptions=chart_descriptions,
                image_data=images,
            )
        )

    return preread


def process_xlsx(file_path: str, llm: LLMClient) -> ExtractionResult:
    """Extract content from an XLSX workbook.

    Phase 1 (sequential): pre-reads all sheet data from openpyxl
    into plain Python structures, then closes both workbooks.

    Phase 2 (parallel): processes each sheet concurrently —
    serializes cells to markdown, describes images via LLM
    vision using pre-loaded bytes. Any sheet failure fails
    the entire file.

    Params:
        file_path: Absolute path to the XLSX file
        llm: LLMClient instance

    Returns:
        ExtractionResult with one PageResult per worksheet

    Example:
        >>> result = process_xlsx("/data/report.xlsx", llm)
        >>> result.filetype
        "xlsx"
    """
    workbook_label = Path(file_path).name
    workbook, cached_workbook = _open_workbooks(file_path)

    try:
        preread = _preread_all_sheets(workbook, cached_workbook)
    finally:
        workbook.close()
        cached_workbook.close()

    has_any_images = any(sd.image_data for sd in preread)
    visual_prompt: dict[str, Any] | None = None
    if has_any_images:
        visual_prompt = load_prompt("visual_extraction", _PROMPTS_DIR)

    cache_state = _ImageDescriptionState()
    total_sheets = len(preread)
    page_workers = get_extraction_page_workers()

    results: list[PageResult | None] = [None] * total_sheets
    with ThreadPoolExecutor(max_workers=page_workers) as pool:
        futures = {
            pool.submit(
                _process_single_sheet,
                sd,
                llm,
                visual_prompt,
                workbook_label,
                total_sheets,
                cache_state,
            ): idx
            for idx, sd in enumerate(preread)
        }
        for future in as_completed(futures):
            results[futures[future]] = future.result()

    pages = [p for p in results if p is not None]

    return ExtractionResult(
        file_path=file_path,
        filetype="xlsx",
        pages=pages,
        total_pages=len(pages),
    )
